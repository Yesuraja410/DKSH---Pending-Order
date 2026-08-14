# -*- coding: utf-8 -*-
import io
import os
import re
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
import streamlit as st
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Page Config
st.set_page_config(
    page_title="DKSH Pending Order Report Automator",
    page_icon="📧",
    layout="wide",
)

# Account IDs supported for filtering
ACCOUNT_IDS = ["AAFHU", "AAFHB"]

# Constants for Seller Emails
SELLER_EMAILS = {
    "AAFHU": [
        "seller-aafhu-th-hadalabo@graas.ai",
        "seller-aafhu-th-oatside@graas.ai",
        "seller-aafhu-th-lamy@graas.ai",
        "seller-aafhu-th-carglo@graas.ai",
        "seller-aafhu-th-bricks@graas.ai",
        "seller-aafhu-th-energizer@graas.ai"
    ],
    "AAFHB": [
        "seller-aafhb-th-colosure@graas.ai",
        "seller-aafhb-th-hiruscar@graas.ai",
        "seller-aafhb-th-nizoral@graas.ai",
        "seller-aafhb-th-lactacyd@graas.ai",
        "seller-aafhb-th-smithnephew@graas.ai",
        "seller-aafhb-th-aquamaris@graas.ai"
    ]
}

# Column Synonyms for Automatic Mapping.
# NOTE: Lists are ordered most-specific-first — auto_map_headers() tries every
# synonym across ALL headers before moving to the next (lower-priority) synonym.
# This matters because real DKSH exports often have several similarly-named
# columns (e.g. order_status / payment_status / item_status / sla_status, or
# order_id / order_number) — a generic word like "status" must never be allowed
# to win over the exact "payment_status" column just because it appears first
# in the file.
COLUMN_SYNONYMS = {
    "orderNumber": ['order number', 'order_number', 'order no', 'order no.', 'orderno', 'document number', 'so number', 'order id', 'order_id', 'order ID #'],
    "paymentStatus": ['payment status', 'payment_status', 'paymentstatus', 'pay status', 'pay_status'],
    "paymentMethod": ['payment method', 'payment_method', 'payment methods', 'payment_methods', 'payment mode', 'paymentmode', 'paymenttype', 'payment type', 'pay method'],
    "nickname": ['nickname', 'seller nickname', 'store nickname', 'seller_nickname', 'nickname seller', 'seller id', 'store', 'shop', 'brand', 'channel'],
    "orderDate": ['ordered date', 'ordered_date', 'order date', 'order_date', 'orderdate', 'order creation date', 'created date', 'date created', 'createdtime', 'created time', 'date']
}

# Helper Functions
def clean_str(val):
    if pd.isna(val):
        return ""
    return str(val).strip()

def normalize_val(val):
    s = clean_str(val).lower()
    # Strip brackets, underscores, dashes, single quotes, and double quotes
    s = s.replace('[', '').replace(']', '').replace('_', ' ').replace('-', ' ').replace('"', ' ').replace("'", ' ')
    return " ".join(s.split())

def parse_date(date_str):
    s = clean_str(date_str)
    if not s:
        return "Unknown Date"
    try:
        dt = datetime.strptime(s.split()[0], "%d/%m/%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        try:
            dt = pd.to_datetime(s)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return s

def get_day_suffix(day):
    if 11 <= day <= 13:
        return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

def get_display_date(date_str):
    if not_str := clean_str(date_str):
        if not_str == 'Unknown Date':
            return not_str
        try:
            dt = datetime.strptime(not_str, "%Y-%m-%d")
            return dt.strftime("%d-%b")
        except Exception:
            return not_str
    return "Unknown Date"

def auto_map_headers(headers):
    """Map each logical field (orderNumber, paymentStatus, ...) to the best
    matching real column name.

    Matching is done by SYNONYM PRIORITY, not by column order: for a given
    field, every synonym in the list is tried (in order) across all headers
    for an exact match before we ever fall back to the next, less specific
    synonym or to a loose "contains" match. This avoids a common bug where a
    generic/ambiguous column (e.g. "order_status") that happens to appear
    earlier in the sheet gets mistakenly picked over the real, exact-match
    column (e.g. "payment_status") purely because of column ordering.
    """
    mapping = {k: "" for k in COLUMN_SYNONYMS}
    normalized_headers = [(h, normalize_val(h)) for h in headers]

    for key, synonyms in COLUMN_SYNONYMS.items():
        # 1) Exact match, tried in synonym priority order (most specific first)
        for syn in synonyms:
            match = next((orig for orig, norm in normalized_headers if norm == syn), None)
            if match:
                mapping[key] = match
                break
        if mapping[key]:
            continue
        # 2) Fallback: loose contains-match, also tried in priority order.
        # Skip very short/generic tokens here to reduce false positives.
        for syn in synonyms:
            if len(syn) <= 3:
                continue
            match = next((orig for orig, norm in normalized_headers if syn in norm or norm in syn), None)
            if match:
                mapping[key] = match
                break
    return mapping

def get_google_sheet_csv_url(share_url):
    id_match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", share_url)
    if not id_match:
        return None
    sheet_id = id_match.group(1)
    gid_match = re.search(r"gid=([0-9]+)", share_url)
    gid = gid_match.group(1) if gid_match else "0"
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

# A real browser User-Agent header. Google's export endpoint sometimes blocks
# or blanks the response for default python-requests / bot-like user agents
# even when a sheet is shared publicly.
GOOGLE_SHEET_REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
}

# SMTP Load/Save Configuration
local_config = {}
try:
    if os.path.exists("config.json"):
        with open("config.json", "r") as f:
            cfg = json.load(f)
            local_config = cfg.get("smtp_config", {})
except Exception:
    pass

smtp_defaults = local_config

# Main Title UI
st.title("📧 DKSH Pending Order Report Automator & Emailer")
st.write("Process your Pending orders (NOT_INITIATED status), aggregate a formatted 2D Pivot grid, and email reports directly to your sellers.")

# Account Selection - placed at the very top, before any data is uploaded,
# so the chosen account drives the output file name / email recipients from
# the start rather than depending on what's auto-detected inside the file.
st.markdown("### 🎯 Select Account")
selected_merchant = st.selectbox(
    "Which account is this report for?",
    ["All Accounts"] + list(ACCOUNT_IDS),
    index=0,
    help="Choose AAFHU or AAFHB before uploading your data. This drives the output file name, email recipients, and (if the uploaded file has a merchant/account column) row filtering."
)
st.markdown("---")

# Layout: Split Config and Data
col_left, col_right = st.columns([1, 2.2])

# Left Column: Config Panel
with col_left:
    st.markdown("### ⚙️ SMTP Email Configuration")
    c_host = st.text_input("SMTP Server Host", value=smtp_defaults.get("host", "smtp.office365.com"))
    c_port = st.text_input("SMTP Port", value=str(smtp_defaults.get("port", 587)))
    c_user = st.text_input("SMTP Username", value=smtp_defaults.get("user", ""))
    c_pass = st.text_input("SMTP Password", type="password", value=smtp_defaults.get("password", ""))
    c_sender = st.text_input("Sender Email Address", value=smtp_defaults.get("sender_email", smtp_defaults.get("user", "")))
    c_tls = st.checkbox("Use TLS", value=smtp_defaults.get("use_tls", True))
    
    if st.button("Test Connection"):
        with st.spinner("Connecting..."):
            try:
                host = c_host
                port = int(c_port) if c_port.isdigit() else 587
                if c_tls:
                    server = smtplib.SMTP(host, port, timeout=10)
                    server.starttls()
                else:
                    server = smtplib.SMTP_SSL(host, port, timeout=10)
                server.login(c_user, c_pass)
                server.close()
                st.success("✅ SMTP connection verified successfully!")
                
                # Save config
                cfg_data = {"smtp_config": {
                    "host": c_host,
                    "port": port,
                    "user": c_user,
                    "password": c_pass,
                    "sender_email": c_sender,
                    "use_tls": c_tls
                }}
                with open("config.json", "w") as f:
                    json.dump(cfg_data, f, indent=4)
            except Exception as e:
                st.error(f"❌ Connection failed: {e}")

# Save SMTP settings dictionary
smtp_config = {
    "host": c_host,
    "port": int(c_port) if c_port.isdigit() else 587,
    "user": c_user,
    "password": c_pass,
    "sender_email": c_sender,
    "use_tls": c_tls
}

# Right Column: Data Source and Processing Dashboard
with col_right:
    st.markdown("### 📊 Data Import & Processing")
    import_type = st.radio("Data Source Type", ["Google Sheets Link", "CSV/XLSX File Upload"], horizontal=True)
    
    df_raw = None
    file_name = ""
    
    if import_type == "Google Sheets Link":
        sheet_url = st.text_input("Paste Public Google Sheet URL", value="", placeholder="https://docs.google.com/spreadsheets/d/...")
        st.caption("Tip: the URL must point to the exact tab you want (open that tab first so the `gid=` in the address bar matches it). The spreadsheet must be shared as **Anyone with the link → Viewer** (Share button in Google Sheets) — 'Restricted' sharing will fail even if you're logged in yourself.")
        if sheet_url:
            with st.spinner("Fetching Google Sheet..."):
                try:
                    csv_url = get_google_sheet_csv_url(sheet_url)
                    if not csv_url:
                        st.error("Invalid Google Sheet URL format. Expected something like https://docs.google.com/spreadsheets/d/<ID>/edit#gid=<TAB_ID>")
                    else:
                        response = requests.get(csv_url, headers=GOOGLE_SHEET_REQUEST_HEADERS, timeout=20, allow_redirects=True)
                        body = response.text.strip()
                        content_type = response.headers.get("Content-Type", "").lower()

                        if response.status_code != 200:
                            st.error(f"Failed to fetch Google Sheet (HTTP {response.status_code}). Verify sharing permissions and that the link is correct.")
                        elif not body:
                            st.error(
                                "Google Sheet returned no data (empty response). This almost always means the sheet "
                                "is not publicly viewable. In Google Sheets, click **Share** → set **General access** "
                                "to **'Anyone with the link'** (role: Viewer) → Done, then try again."
                            )
                        elif body.lstrip().lower().startswith("<") or "text/html" in content_type:
                            st.error(
                                "Google returned a sign-in / permission page instead of your data. Please set the "
                                "sheet's sharing to **'Anyone with the link'** (Share button in Google Sheets), then try again. "
                                "If it's already shared that way, double check the `gid=` in your URL matches the tab with your data."
                            )
                        else:
                            # Read raw CSV using text strings to avoid large ID scientific notation crashes
                            df_raw = pd.read_csv(io.StringIO(response.text), dtype=str).fillna('')
                            if df_raw.empty or len(df_raw.columns) == 0:
                                st.error("The fetched tab appears to be empty. Double-check the `gid=` in your URL points to the correct tab (the one with your order data).")
                                df_raw = None
                            else:
                                file_name = "Google_Sheet_Data"
                                st.success(f"✅ Successfully fetched Google Sheet data! ({len(df_raw)} rows, {len(df_raw.columns)} columns)")
                except Exception as e:
                    st.error(f"Error loading Google Sheet: {e}")
    else:
        uploaded_file = st.file_uploader("Upload Order CSV/XLSX Report", type=["csv", "xlsx", "xls"])
        if uploaded_file:
            with st.spinner("Reading uploaded file..."):
                try:
                    file_name = uploaded_file.name
                    if file_name.lower().endswith(".csv"):
                        df_raw = pd.read_csv(uploaded_file, dtype=str).fillna('')
                    else:
                        # Open Excel and force string cells to prevent decimal rounding or float scientific conversions
                        wb_temp = openpyxl.load_workbook(uploaded_file, read_only=True)
                        ws_temp = wb_temp[wb_temp.sheetnames[0]]
                        data = []
                        headers = []
                        for idx, row in enumerate(ws_temp.iter_rows(values_only=True)):
                            if idx == 0:
                                headers = [str(cell).strip() if cell is not None else f"Column_{i}" for i, cell in enumerate(row)]
                            else:
                                row_dict = {}
                                for c_idx, cell in enumerate(row):
                                    header = headers[c_idx] if c_idx < len(headers) else f"Column_{c_idx}"
                                    row_dict[header] = str(cell).strip() if cell is not None else ""
                                data.append(row_dict)
                        df_raw = pd.DataFrame(data)
                    st.success("✅ Successfully loaded file data!")
                except Exception as e:
                    st.error(f"Error reading file: {e}")
                    
    # Processing Section
    if df_raw is not None:
        st.markdown("---")
        
        # Display mapping configuration
        headers_list = list(df_raw.columns)
        auto_map = auto_map_headers(headers_list)
        
        with st.expander("🛠️ Columns Mapping Details", expanded=False):
            col_map_order = st.selectbox("Order Number Column", headers_list, index=headers_list.index(auto_map["orderNumber"]) if auto_map["orderNumber"] in headers_list else 0)
            col_map_status = st.selectbox("Payment Status Column", headers_list, index=headers_list.index(auto_map["paymentStatus"]) if auto_map["paymentStatus"] in headers_list else 0)
            col_map_method = st.selectbox("Payment Method Column", headers_list, index=headers_list.index(auto_map["paymentMethod"]) if auto_map["paymentMethod"] in headers_list else 0)
            col_map_nickname = st.selectbox("Store Nickname Column", headers_list, index=headers_list.index(auto_map["nickname"]) if auto_map["nickname"] in headers_list else 0)
            col_map_date = st.selectbox("Order Date Column", headers_list, index=headers_list.index(auto_map["orderDate"]) if auto_map["orderDate"] in headers_list else 0)
        
        mapping = {
            "orderNumber": col_map_order,
            "paymentStatus": col_map_status,
            "paymentMethod": col_map_method,
            "nickname": col_map_nickname,
            "orderDate": col_map_date
        }
        
        # Detect the Account/Merchant ID column in the uploaded data so we can
        # filter rows to match the account chosen at the top of the page.
        # (The account itself is already selected above, before upload.)
        ACCOUNT_COL_CANDIDATES = [
            'merchant_id', 'merchant', 'account_id', 'account', 'seller_id',
            'seller_account', 'account_no', 'account number', 'merchantid',
            'accountid', 'account name', 'account_name'
        ]
        merchant_col = None
        for c in df_raw.columns:
            if c.strip().lower() in ACCOUNT_COL_CANDIDATES:
                merchant_col = c
                break

        # Fallback: scan every column for one that actually contains AAFHU/AAFHB values
        if merchant_col is None:
            for c in df_raw.columns:
                sample_vals = df_raw[c].astype(str).str.upper()
                if sample_vals.str.contains('AAFHU|AAFHB', regex=True, na=False).any():
                    merchant_col = c
                    break

        st.info(f"📬 Processing report for account: **{selected_merchant}**")
        if selected_merchant != "All Accounts" and merchant_col is None:
            st.warning(f"⚠️ Could not find an Account/Merchant ID column containing '{selected_merchant}' in the dataset. All rows will be processed without account filtering.")

        # Core Processor Trigger
        # Run exclusion filter and deduplication
        with st.spinner("Processing & Formatting Report..."):
            # Exclude orders
            df_merch = df_raw.copy()
            if selected_merchant != "All Accounts" and merchant_col is not None:
                df_merch = df_raw[df_raw[merchant_col].astype(str).str.upper().str.contains(selected_merchant.upper(), na=False)]

            filtered_rows = []
            excluded_count = 0

            for idx, row in df_merch.iterrows():
                status = normalize_val(row[mapping["paymentStatus"]]) if mapping["paymentStatus"] in df_merch.columns else ""
                method = normalize_val(row[mapping["paymentMethod"]]) if mapping["paymentMethod"] in df_merch.columns else ""

                # Exclude if payment_status is NOT_INITIATED and payment_method is NOT COD
                is_excluded_status = (status == 'not initiated')
                is_cod = (method == 'cod' or 'cod' in method or ('cash' in method and 'delivery' in method))

                if is_excluded_status and not is_cod:
                    excluded_count += 1
                else:
                    filtered_rows.append(row)
                    
            df_filtered = pd.DataFrame(filtered_rows)
            
            if df_filtered.empty:
                st.error("No orders remain after applying the filter criteria.")
            else:
                df_filtered['clean_date'] = df_filtered[mapping["orderDate"]].apply(parse_date)
                
                # Deduplicate by Order Number for Pivot Summary
                unique_orders = df_filtered.drop_duplicates(subset=[mapping["orderNumber"]], keep='first')
                unique_order_count = len(unique_orders)
                
                # Metrics Summary Cards
                st.markdown("#### 📈 Processing Summary Metrics")
                card1, card2, card3, card4 = st.columns(4)
                card1.metric("Original Rows", len(df_merch))
                card2.metric("Excluded Rows", f"-{excluded_count}")
                card3.metric("Kept Line Items", len(df_filtered))
                card4.metric("Unique Kept Orders", unique_order_count)
                
                # Generate Pivot Summary DataFrame for rendering in App
                pivot_df = unique_orders.groupby([mapping["nickname"], 'clean_date']).size().reset_index(name='order_count')
                dates_in_sheet = sorted([d for d in pivot_df['clean_date'].unique() if d != 'Unknown Date'])
                
                # Build pivot grid
                create_pivot = pivot_df.pivot(index=mapping["nickname"], columns='clean_date', values='order_count').fillna(0).astype(int)
                create_pivot_display = create_pivot.copy()
                # Rename columns for display
                create_pivot_display.columns = [get_display_date(c) for c in create_pivot_display.columns]
                create_pivot_display['Grand Total'] = create_pivot_display.sum(axis=1)
                create_pivot_display.loc['Grand Total'] = create_pivot_display.sum(axis=0)
                
                st.markdown("#### 📊 Order Pivot Summary View")
                # Highlight rows/cells in UI
                st.dataframe(create_pivot_display.style.format(lambda x: int(x) if x != 0 else "-"))
                
                # Build Styled Workbook Excel Report
                wb = openpyxl.Workbook()
                ws_pivot = wb.active
                ws_pivot.title = "Summary"
                ws_pivot.views.sheetView[0].showGridLines = True

                ws_details = wb.create_sheet(title="Data")
                ws_details.views.sheetView[0].showGridLines = True
                
                # Layout formatting openpyxl styles
                font_title = Font(name='Calibri', size=11, bold=True, color='000000')
                fill_title = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
                font_header = Font(name='Calibri', size=10, bold=True, color='000000')
                
                border_black_thin = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                border_grand_total = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='double', color='000000')
                )
                fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                
                # Write Merged Title Row
                num_cols = len(dates_in_sheet) + 2
                ws_pivot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
                
                for col_num in range(1, num_cols + 1):
                    cell = ws_pivot.cell(row=1, column=col_num)
                    cell.fill = fill_title
                    cell.border = border_black_thin
                    
                today_dt = datetime.now()
                formatted_today = f"{today_dt.day}{get_day_suffix(today_dt.day)} {today_dt.strftime('%b')} -{today_dt.year}"
                
                title_cell = ws_pivot.cell(row=1, column=1)
                title_cell.value = f"Pending Order Report - {formatted_today}"
                title_cell.font = font_title
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws_pivot.row_dimensions[1].height = 20
                
                # Write Headers on Row 2
                headers = ['Store / Nickname'] + [get_display_date(d) for d in dates_in_sheet] + ['Grand Total']
                ws_pivot.append(headers)
                ws_pivot.row_dimensions[2].height = 18
                
                for col_num in range(1, num_cols + 1):
                    cell = ws_pivot.cell(row=2, column=col_num)
                    cell.font = font_header
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border_black_thin
                    
                # Write Data Rows
                row_idx = 3
                col_totals = [0] * len(dates_in_sheet)
                overall_total = 0
                max_date_str = dates_in_sheet[-1] if dates_in_sheet else ""
                
                for nickname in sorted(create_pivot.index):
                    row_cells = [nickname]
                    row_total = 0
                    for d_idx, date_str in enumerate(dates_in_sheet):
                        count = int(create_pivot.loc[nickname, date_str])
                        if count > 0:
                            row_cells.append(count)
                            row_total += count
                            col_totals[d_idx] += count
                        else:
                            row_cells.append("")
                    row_cells.append(row_total)
                    overall_total += row_total
                    
                    ws_pivot.append(row_cells)
                    ws_pivot.row_dimensions[row_idx].height = 16
                    
                    for col_num in range(1, num_cols + 1):
                        cell = ws_pivot.cell(row=row_idx, column=col_num)
                        cell.font = Font(name='Calibri', size=10, bold=(col_num == num_cols))
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_black_thin
                        
                        if 2 <= col_num <= num_cols - 1:
                            date_iso = dates_in_sheet[col_num - 2]
                            cell_val = cell.value
                            if date_iso < max_date_str and cell_val != "" and cell_val is not None:
                                cell.fill = fill_yellow
                    row_idx += 1
                    
                # Write Grand Total Row
                grand_total_row = ['Grand Total'] + col_totals + [overall_total]
                ws_pivot.append(grand_total_row)
                ws_pivot.row_dimensions[row_idx].height = 18
                
                for col_num in range(1, num_cols + 1):
                    cell = ws_pivot.cell(row=row_idx, column=col_num)
                    cell.font = Font(name='Calibri', size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border_grand_total
                    
                ws_pivot.column_dimensions['A'].width = 30
                for c in range(2, num_cols + 1):
                    ws_pivot.column_dimensions[get_column_letter(c)].width = 11
                    
                # Write Details sheet (keeping all columns)
                original_cols = list(df_raw.columns)
                ws_details.append(original_cols)
                ws_details.row_dimensions[1].height = 24
                
                header_fill_blue = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
                font_header_details = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
                
                for col_num in range(1, len(original_cols) + 1):
                    cell = ws_details.cell(row=1, column=col_num)
                    cell.fill = header_fill_blue
                    cell.font = font_header_details
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = border_black_thin
                    
                det_row_idx = 2
                for idx, row_series in df_filtered.iterrows():
                    row_data = [row_series[c] for c in original_cols]
                    ws_details.append(row_data)
                    ws_details.row_dimensions[det_row_idx].height = 18
                    
                    for col_num in range(1, len(original_cols) + 1):
                        cell = ws_details.cell(row=det_row_idx, column=col_num)
                        cell.font = Font(name='Calibri', size=10)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border_black_thin
                        
                        col_name = original_cols[col_num - 1]
                        if col_name == mapping["orderNumber"]:
                            cell.value = str(row_series[col_name])
                            cell.number_format = '@'
                    det_row_idx += 1
                    
                # Auto-fit widths details sheet
                for col_num in range(1, len(original_cols) + 1):
                    col_letter = get_column_letter(col_num)
                    max_len = len(original_cols[col_num - 1])
                    for r_num in range(2, det_row_idx):
                        val_str = str(ws_details.cell(row=r_num, column=col_num).value or '')
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    ws_details.column_dimensions[col_letter].width = min(max_len + 4, 40)
                    
                # Compile bytes output
                out_excel = io.BytesIO()
                wb.save(out_excel)
                excel_bytes = out_excel.getvalue()
                
                st.markdown("---")
                st.markdown("### 📤 Download & Forward Actions")
                
                action_col1, action_col2 = st.columns(2)
                
                with action_col1:
                    # Download File Button
                    st.download_button(
                        label="💾 Download Excel Report",
                        data=excel_bytes,
                        file_name=f"Pending_Order_Report_{selected_merchant.replace(' ', '_')}_{today_dt.strftime('%d%b_%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    
                with action_col2:
                    # Email Sending Section
                    if selected_merchant in SELLER_EMAILS:
                        recipients = SELLER_EMAILS[selected_merchant]
                        st.info(f"📬 Selected Account: **{selected_merchant}**\n\nRecipients:\n" + "\n".join([f"- `{r}`" for r in recipients]))
                        
                        # Validate SMTP has values
                        if not c_host or not c_user or not c_pass:
                            st.warning("⚠️ Configure and verify your SMTP credentials in the left panel to enable email sending.")
                            st.button("✉️ Send via Email", disabled=True, use_container_width=True)
                        else:
                            if st.button("✉️ Send via Email to Sellers", use_container_width=True):
                                with st.spinner("Compiling and sending email..."):
                                    # Send Email
                                    msg = MIMEMultipart()
                                    msg['From'] = smtp_config['sender_email']
                                    msg['To'] = ", ".join(recipients)
                                    msg['Subject'] = f"Pending Order Report - {selected_merchant} - {formatted_today}"
                                    
                                    body_text = (
                                        f"Hello Sellers,\n\n"
                                        f"Please find attached the Pending Order Report for {selected_merchant} "
                                        f"generated on {formatted_today}.\n\n"
                                        f"Total unique pending orders: {unique_order_count}\n\n"
                                        f"Best Regards,\n"
                                        f"DKSH Order Management System Automator"
                                    )
                                    msg.attach(MIMEText(body_text, 'plain'))
                                    
                                    part = MIMEBase('application', 'octet-stream')
                                    part.set_payload(excel_bytes)
                                    encoders.encode_base64(part)
                                    part.add_header(
                                        'Content-Disposition',
                                        f'attachment; filename="Pending_Order_Report_{selected_merchant}_{today_dt.strftime("%d%b_%Y")}.xlsx"'
                                    )
                                    msg.attach(part)
                                    
                                    try:
                                        if smtp_config['use_tls']:
                                            server = smtplib.SMTP(smtp_config['host'], smtp_config['port'], timeout=15)
                                            server.starttls()
                                        else:
                                            server = smtplib.SMTP_SSL(smtp_config['host'], smtp_config['port'], timeout=15)
                                            
                                        server.login(smtp_config['user'], smtp_config['password'])
                                        server.sendmail(smtp_config['sender_email'], recipients, msg.as_string())
                                        server.close()
                                        st.success(f"🎉 Report successfully emailed to the {selected_merchant} seller group!")
                                    except Exception as e:
                                        st.error(f"❌ Failed to send email: {e}")
                    else:
                        st.info("💡 Please select a specific Merchant Account (AAFHU or AAFHB) from the dropdown above to enable direct seller email forwarding.")
                        st.button("✉️ Send via Email", disabled=True, use_container_width=True)
