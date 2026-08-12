import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv_path = r"C:\Users\Yesuraja\.gemini\antigravity\scratch\Pending order report\temp_reconciliation\multiple_channels-23-Jul-2026.csv"
output_path = r"C:\Users\Yesuraja\.gemini\antigravity\scratch\pending-order-dashboard\Pending_Order_Report_Output.xlsx"

print("Reading input CSV with string datatypes...")
df = pd.read_csv(csv_path, dtype=str).fillna('')

# Map columns
col_order_no = 'order_number'
col_nickname = 'nickname'
col_status = 'payment_status'
col_method = 'payment_methods'
col_date = 'ordered_date'

# Clean strings
def clean_str(val):
    return str(val).strip()

def normalize_val(val):
    s = clean_str(val).lower()
    s = s.replace('[', '').replace(']', '').replace('_', ' ').replace('-', ' ')
    return " ".join(s.split())

# 1. Exclude orders: Payment Status is NOT_INITIATED or Pending and Payment Method is not COD
filtered_rows = []
for idx, row in df.iterrows():
    status = normalize_val(row[col_status])
    method = normalize_val(row[col_method])
    is_excluded_status = status == 'not initiated' or status == 'pending'
    is_cod = 'cod' in method or ('cash' in method and 'delivery' in method)
    if not (is_excluded_status and not is_cod):
        filtered_rows.append(row)

df_filtered = pd.DataFrame(filtered_rows)

# Standardize dates for Pivot aggregation
def parse_date(date_str):
    s = clean_str(date_str)
    if not s:
        return "Unknown Date"
    try:
        dt = datetime.strptime(s.split()[0], "%d/%m/%Y")
        return dt.strftime("%Y-%m-%d")
    except Exception:
        try:
            dt = pd.to_datetime(s)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return s

df_filtered['clean_date'] = df_filtered[col_date].apply(parse_date)

# 2. For the PIVOT table, we deduplicate by Order Number
unique_orders = df_filtered.drop_duplicates(subset=[col_order_no], keep='first')

# 3. Generate Pivot Table Data from deduplicated orders
pivot_df = unique_orders.groupby([col_nickname, 'clean_date']).size().reset_index(name='order_count')

# Pivot chronologically sorted
dates_in_sheet = sorted([d for d in pivot_df['clean_date'].unique() if d != 'Unknown Date'])
max_date_str = dates_in_sheet[-1] if dates_in_sheet else ""

create_pivot = pivot_df.pivot(index=col_nickname, columns='clean_date', values='order_count').fillna(0).astype(int)

# Helper to format display date (e.g. 2026-07-22 -> 22-Jul)
def get_display_date(date_str):
    if not_str := clean_str(date_str):
        if not_str == 'Unknown Date':
            return not_str
        try:
            dt = datetime.strptime(not_str, "%Y-%m-%d")
            return dt.strftime("%d-%b")
        except Exception:
            return not_str
    return "Unknown Date"

# Compute suffix for header date (e.g., 12 -> 12th)
def get_day_suffix(day):
    if 11 <= day <= 13:
        return 'th'
    return {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')

today = datetime.now()
day_suffix = get_day_suffix(today.day)
formatted_today = f"{today.day}{day_suffix} {today.strftime('%b')} -{today.year}"

# Build Workbook
wb = openpyxl.Workbook()

# Sheet 1: Pivot Summary
ws_pivot = wb.active
ws_pivot.title = "Pivot Summary"
ws_pivot.views.sheetView[0].showGridLines = True

# Sheet 2: Cleaned Details (holds all original columns and rows, no duplicates removed here)
ws_details = wb.create_sheet(title="Cleaned Details")
ws_details.views.sheetView[0].showGridLines = True

# Styling helpers (mockup aligned)
font_title = Font(name='Calibri', size=11, bold=True, color='000000')
fill_title = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid") # light blue
font_header = Font(name='Calibri', size=10, bold=True, color='000000')

# Borders must be black thin borders as shown in mockup
border_black_thin = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
border_grand_total = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)

fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # bright yellow

# Write Title Row in Pivot
num_cols = len(dates_in_sheet) + 2 # Nickname + dates + total
ws_pivot.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

# Apply styling/border to all merged header cells
for col_num in range(1, num_cols + 1):
    cell = ws_pivot.cell(row=1, column=col_num)
    cell.fill = fill_title
    cell.border = border_black_thin

title_cell = ws_pivot.cell(row=1, column=1)
title_cell.value = f"Pending Order Report - {formatted_today}"
title_cell.font = font_title
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws_pivot.row_dimensions[1].height = 20

# Write headers (directly on row 2, no spacer row)
headers = ['Store / Nickname'] + [get_display_date(d) for d in dates_in_sheet] + ['Grand Total']
ws_pivot.append(headers)
ws_pivot.row_dimensions[2].height = 18

# Style headers
for col_num in range(1, num_cols + 1):
    cell = ws_pivot.cell(row=2, column=col_num)
    cell.font = font_header
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_black_thin

# Write Pivot rows
row_idx = 3
col_totals = [0] * len(dates_in_sheet)
overall_total = 0

for nickname in sorted(create_pivot.index):
    row_cells = [nickname]
    row_total = 0
    
    # Calculate counts and totals
    for d_idx, date_str in enumerate(dates_in_sheet):
        count = int(create_pivot.loc[nickname, date_str])
        if count > 0:
            row_cells.append(count)
            row_total += count
            col_totals[d_idx] += count
        else:
            row_cells.append("") # blank if 0
            
    row_cells.append(row_total)
    overall_total += row_total
    
    ws_pivot.append(row_cells)
    ws_pivot.row_dimensions[row_idx].height = 16
    
    # Style row cells
    for col_num in range(1, num_cols + 1):
        cell = ws_pivot.cell(row=row_idx, column=col_num)
        cell.font = Font(name='Calibri', size=10, bold=(col_num == num_cols))
        cell.alignment = Alignment(horizontal="center", vertical="center") # Center aligned nicknames & values
        cell.border = border_black_thin
        
        # Highlight yellow for past date cells containing values
        if 2 <= col_num <= num_cols - 1:
            date_iso = dates_in_sheet[col_num - 2]
            cell_val = cell.value
            if date_iso < max_date_str and cell_val != "" and cell_val is not None:
                cell.fill = fill_yellow
                
    row_idx += 1

# Write Grand Total Row
grand_total_row = ['Grand Total'] + col_totals + [overall_total]
ws_pivot.append(grand_total_row)
ws_pivot.row_dimensions[row_idx].height = 18

# Style Grand Total Row
for col_num in range(1, num_cols + 1):
    cell = ws_pivot.cell(row=row_idx, column=col_num)
    cell.font = Font(name='Calibri', size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_grand_total

# Set compact column widths for Pivot sheet
ws_pivot.column_dimensions['A'].width = 30
for c in range(2, num_cols + 1):
    col_letter = get_column_letter(c)
    ws_pivot.column_dimensions[col_letter].width = 11

# --- Write Details Sheet ---
# Keep all original columns from CSV
original_cols = list(df.columns)
ws_details.append(original_cols)
ws_details.row_dimensions[1].height = 24

# Style headers for details
header_fill_blue = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") # bold blue background
font_header_details = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

for col_num in range(1, len(original_cols) + 1):
    cell = ws_details.cell(row=1, column=col_num)
    cell.fill = header_fill_blue
    cell.font = font_header_details
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_black_thin

# Write details data rows
det_row_idx = 2
for idx, row in df_filtered.iterrows():
    row_data = [row[c] for c in original_cols]
    ws_details.append(row_data)
    ws_details.row_dimensions[det_row_idx].height = 18
    
    # Style data rows (center aligned, force text format for order numbers to prevent scientific format)
    for col_num in range(1, len(original_cols) + 1):
        cell = ws_details.cell(row=det_row_idx, column=col_num)
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_black_thin
        
        col_name = original_cols[col_num - 1]
        if col_name == col_order_no:
            # Force exact value as text string format
            cell.value = str(row[col_name])
            cell.number_format = '@'
            
    det_row_idx += 1

# Auto-fit details widths (capped at 40 max width)
for col_num in range(1, len(original_cols) + 1):
    col_letter = get_column_letter(col_num)
    max_len = len(original_cols[col_num - 1])
    for row in range(2, det_row_idx):
        val = str(ws_details.cell(row=row, column=col_num).value or '')
        if len(val) > max_len:
            max_len = len(val)
    ws_details.column_dimensions[col_letter].width = min(max_len + 4, 40)

# Save workbook
wb.save(output_path)
print(f"SUCCESS: Unified styled spreadsheet saved to {output_path}")
